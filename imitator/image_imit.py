from PIL import Image, ImageDraw, ImageFont
from altair import renderers
from matplotlib.colors import LightSource
import matplotlib.font_manager as fm
import math

import streamlit as st
import numpy as np
import matplotlib.pyplot as plt
import io
import random
import cv2
import os
import openpyxl
import pythreejs as p3
import base64
from IPython.display import display
from io import BytesIO
import moderngl
from pyvista import Plotter, StructuredGrid
import pyvista as pv
from matplotlib.colors import LightSource

import imrotate

database_path = "ДеталиПоПлануДляРазрешенныхЗаказов.xlsx"

def get_random_cropped_texture_image(size):
    # Получаем список всех файлов в папке ./textures
    textures_folder = './textures'
    files = os.listdir(textures_folder)

    # Выбираем случайный файл
    random_file = random.choice(files)
    image_path = os.path.join(textures_folder, random_file)

    # Открываем изображение
    image = Image.open(image_path)
    if image.mode == 'L':  # 'L' означает градации серого
        image = image.convert('RGB')  # Преобразование в RGB

    # Поворачиваем изображение на случайный угол
    angle = random.uniform(0, 360)
    image = imrotate.make_rotate(image, angle) # Поворот и обрезка черных областей

    # Масштабируем изображение в случайных пределах от 0.5 до 2
    scale = random.uniform(0.5, 2)
    new_size = (int(image.width * scale), int(image.height * scale))
    image = image.resize(new_size, Image.Resampling.LANCZOS)

    # Определяем случайную позицию для вырезки
    max_left = max(0, image.width - size[0])
    max_top = max(0, image.height - size[1])
    left = random.randint(0, max_left)
    top = random.randint(0, max_top)
    right = left + size[0]
    bottom = top + size[1]

    # Вырезаем случайную часть заданного размера
    image = image.crop((left, top, right, bottom))

    return image

def generate_linear_transpared_image(size,angle1, angle2):
    h1 = size[0] * math.tan(angle1 * 3.1415 * 180) / 2
    h2 = size[1] * math.tan(angle2 * 3.1415 * 180) / 2
    linear_spacings2 = np.linspace(start=(1 - h1), stop=(1 + h1), num=size[0]) # h
    linear_spacings1 = np.linspace(start=(1 - h2), stop=(1 + h2), num=size[1]) # w
    matrix1 = np.repeat(linear_spacings1[np.newaxis, :], size[0], axis=0)
    matrix2 = np.repeat(linear_spacings2[:, np.newaxis], size[1], axis=1)

    return (matrix1+matrix2)

def create_circle_image(input_image, circle_radius=3):
    # Конвертируем входное изображение в градации серого
    img = input_image.convert("L")
    img_array = np.array(img)

    # Создаем новое изображение для результата
    result_img = Image.new("L", img.size, 127)  # Белый фон
    draw = ImageDraw.Draw(result_img)

    result_img_tmp = Image.new("L", img.size, 127)  # Белый фон
    draw_tmp = ImageDraw.Draw(result_img_tmp)


    k = 1.9 # Коэффициент перекрытия *2
    # Проходим по каждому пикселю изображения
    for y in range(0, img_array.shape[0], 1):
        for x in range(0, img_array.shape[1], 1):
            # Проверяем, находится ли текущая позиция в пределах изображения
            if x < img_array.shape[1] and y < img_array.shape[0]:
                if img_array[y, x] < 100 and result_img_tmp.getpixel((x, y))!=0:
                    draw.ellipse((x - circle_radius, y - circle_radius, x + circle_radius, y + circle_radius), fill=0)
                    draw_tmp.ellipse((x - circle_radius*k, y - circle_radius*k, x + circle_radius*k, y + circle_radius*k), fill=0)
    return result_img

def rotate_symbol_without_black_area(img, angle):
    # Invert colors, rotate, and apply Gaussian blur (similar to original code)
    img = Image.eval(img, lambda x: 255 - x)
    img = img.rotate(angle)
    img = np.array(img)
    img[img == 0] = 127
    img = 255 - img
    img = Image.fromarray(img, mode='L')
    return img

# Функция для генерации изображения текста
def generate_image_with_word(word, font_size, font_path, size, symbol_angle):
    width0 = size[0]*len(word)//3
    height0 = size[1]

    height = max(height0, abs(round(width0 * math.sin(symbol_angle * 3.1415 / 180))))
    width = max(width0, abs(round(height0 * math.cos(symbol_angle * 3.1415 / 180))))

    img = Image.new('L', [width,height], color=127)  # 'L' - grayscale
    draw = ImageDraw.Draw(img)

    font = ImageFont.truetype(font_path, font_size)

    # Calculate text size and position for centering
    #text_width, text_height = draw.textsize(word, font=font)
    text_width, text_height = draw.textbbox((0, 0), word, font=font)[2:4]

    position = ((width - text_width) // 2, (height - text_height) // 2)

    # Draw the word
    draw.text(position, word, fill=0, font=font)

    etalon_img = img

    blur_size1 = 7
    blur_size2 = 7
    if random.choice([0, 1]):
        img = create_circle_image(img, 3.5)
        blur_size1 = 3
        blur_size2 = 3

    img = rotate_symbol_without_black_area(img, symbol_angle)
    etalon_img = rotate_symbol_without_black_area(etalon_img, symbol_angle)

    img = cv2.GaussianBlur(np.array(img), (blur_size1, blur_size2), 0)

    return img, etalon_img

def set_axes_equal(ax):
    """
    Make axes of 3D plot have equal scale so that spheres appear as spheres,
    cubes as cubes, etc.

    Input
      ax: a matplotlib axis, e.g., as output from plt.gca().
    """

    x_limits = ax.get_xlim3d()
    y_limits = ax.get_ylim3d()
    z_limits = ax.get_zlim3d()

    x_range = abs(x_limits[1] - x_limits[0])
    x_middle = np.mean(x_limits)
    y_range = abs(y_limits[1] - y_limits[0])
    y_middle = np.mean(y_limits)
    z_range = abs(z_limits[1] - z_limits[0])
    z_middle = np.mean(z_limits)

    # The plot bounding box is a sphere in the sense of the infinity
    # norm, hence I call half the max range the plot radius.
    plot_radius = 0.5*max([x_range, y_range, z_range])

    ax.set_xlim3d([x_middle - plot_radius, x_middle + plot_radius])
    ax.set_ylim3d([y_middle - plot_radius, y_middle + plot_radius])
    ax.set_zlim3d([z_middle - plot_radius, z_middle + plot_radius])

def crop_render_image(img):
        open_cv_image = np.array(img)
        #open_cv_image = open_cv_image[:, :, ::-1].copy()

        gray = cv2.cvtColor(open_cv_image, cv2.COLOR_BGR2GRAY)
        _, thresh = cv2.threshold(gray, 254, 255, cv2.THRESH_BINARY_INV)
        contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        # Find the largest contour (the rectangle)
        if contours:
            largest_contour = max(contours, key=cv2.contourArea)
            x, y, w, h = cv2.boundingRect(largest_contour)

            # Crop the image
            cropped_img = open_cv_image[y:y+h, x:x+w]

            # Convert back to Pillow format
            cropped_pillow_img = Image.fromarray(cv2.cvtColor(cropped_img, cv2.COLOR_BGR2RGB))
            return cropped_pillow_img
        else:
            return img


def render_surface(pixels, texture_pixels, scale_factor, light_azdeg, light_altdeg, elev, azim, light_power):
    # Получаем изображение текстуры, вдавливаем в нее изображение символов и делаем рендер с различной освещенностью

    pixels = pixels[:, ::-1]
    height, width = pixels.shape

    # Создаем сетку для 3D-объекта
    x = np.linspace(0, width, width)
    y = np.linspace(0, height, height)
    x, y = np.meshgrid(x, y)
    z = (pixels / 255.0) * scale_factor*1  # Нормализуем значение яркости

    # Добавим поверхности случайную шероховатость
    z = z + np.random.normal(0, np.random.uniform(0, 0.05), z.shape)

    # Наклон поверхности
    z = z + generate_linear_transpared_image([height, width], 90-elev, 90-azim)

    # Визуализируем 3D-объект
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')

    ls = LightSource(azdeg=light_azdeg, altdeg=light_altdeg)
    illuminated_surface = ls.shade_rgb(texture_pixels, z, blend_mode="overlay", fraction=light_power)

    # Отображаем поверхность с текстурой и освещением
    surf = ax.plot_surface(x, y, z, rstride=1, cstride=1, shade=True, linewidth=0, antialiased=False, facecolors=illuminated_surface)

    # Установка одинаковых пропорций для осей
    max_range = np.array([x.max()-x.min(), y.max()-y.min(), z.max()-z.min()]).max() / 2.0

    mid_x = (x.max()+x.min()) * 0.5
    mid_y = (y.max()+y.min()) * 0.5
    mid_z = (z.max()+z.min()) * 0.5
    #
    ax.set_xlim(mid_x - max_range, mid_x + max_range)
    ax.set_ylim(mid_y - max_range, mid_y + max_range)
    ax.set_zlim(mid_z - max_range, mid_z + max_range)

    # Убираем оси
    ax.set_axis_off()

    # Устанавливаем угол обзора
    ax.view_init(elev=90, azim=90)

    # Обновляем график перед сохранением
    plt.draw()
    plt.show()

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', dpi=300)
    buf.seek(0)
    img = Image.open(buf).copy()

    buf.close()  # Закрываем буфер
    plt.close(fig)  # Закрываем фигуру после сохранения

    return img


def crop_image(image, crop_percentage):

    # Получаем размеры изображения
    width, height = image.size

    # Вычисляем новые размеры
    new_width = int(width * (1 - crop_percentage / 100))   # Новый размер по ширине
    new_height = int(height * (1 - crop_percentage / 100))  # Новый размер по высоте

    # Вычисляем координаты для обрезки
    left = int(width * (crop_percentage / 200))   # Левый край (половина процента от ширины)
    top = int(height * (crop_percentage / 200))    # Верхний край (половина процента от высоты)
    right = left + new_width                        # Правая граница
    bottom = top + new_height                       # Нижняя граница

    # Обрезаем изображение
    cropped_image = image.crop((left, top, right, bottom))

    return cropped_image

def get_random_cell_from_first_column(filepath=database_path):
    #Открывает файл Excel, выбирает случайную ячейку из первого столбца и возвращает её текстовое значение.
    workbook = openpyxl.load_workbook(filepath, read_only=True)  # read_only for better performance
    sheet = workbook.active  # Get the active sheet
    #Generate random row number
    random_row = random.randint(1, sheet.max_row)
    # Get the cell value
    cell_value = sheet.cell(row=random_row, column=1).value
    return cell_value[1:-1]


def check_image(img):

    width, height = img.size

    # Проверяем размеры
    if width < 20 or height < 20:
        # Создаем новое черное изображение размером 128x128
        img = Image.new("RGB", (128, 128), (0, 0, 0))
        print("warning zero image size")
    return img


def generate(leter, font, scale_factor, symbpl_angle, light_azdeg, light_altdeg, elev, azim, light_power):
    size = [128,128]

    img_2Dletter, etalon = generate_image_with_word(leter, 72, font, size, symbpl_angle)
    height, width = img_2Dletter.shape

    img_texture = get_random_cropped_texture_image([width, height])

    img = render_surface(np.array(img_2Dletter),
                                    np.array(img_texture) / 255.0,
                                    scale_factor,
                                    light_azdeg,
                                    light_altdeg,
                                    90 + elev,
                                    90 + azim,
                                    light_power
                                    )


    img = crop_render_image(img)
    img = crop_image(img, 10)
    return check_image(img), etalon





# Проверочная генерация
# img_2Dletter, etalon = generate_image_with_word(get_random_cell_from_first_column(),72, './fonts/GOST.ttf', size=[128,128], symbol_angle=30)
# height, width = img_2Dletter.shape
# #width, height = img_2Dletter.image.size
# im = plt.imshow(img_2Dletter, cmap='hot')
# plt.colorbar(im, orientation='horizontal')
# plt.show()
#
# img_texture = get_random_cropped_texture_image([width, height])
# #img_texture = np.array(img_texture)
#
# #height, width = img_texture.shape
# im = plt.imshow(img_texture)
# plt.show()
#
# img = render_surface(np.array(img_2Dletter),
#     np.array(img_texture) / 255.0,
#     scale_factor=1,
#     light_azdeg=10,
#     light_altdeg=0,
#     elev=90+0,
#     azim=90-0,
#     light_power=1
# )
#
# plt.show()
#
# img = crop_render_image(img)
# plt.imshow(img)
# plt.show()
